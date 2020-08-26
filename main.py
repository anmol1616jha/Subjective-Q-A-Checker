import language_check
import nltk
import openpyxl
from gensim.summarization import keywords
from nltk.corpus import wordnet
from nltk.stem import WordNetLemmatizer

my_path = "F:\Projects\Subjective-Q-A-Checker\QA Checker\Raw.xlsx" # Give the location of the result excel file

my_wb_obj = openpyxl.load_workbook(my_path)
my_sheet_obj = my_wb_obj.active
my_row = my_sheet_obj.max_row

totalMarks = 10

###########################################
############_Keyword Extraction_###########
###########################################

lemmatizer = WordNetLemmatizer()

sample_answer = open('Sample.txt', 'r')
text = sample_answer.read()

sample_word = keywords(
    text,
    words = 50
).split("\n")

sample_words = set()
for w in sample_word:
    temp = lemmatizer.lemmatize(w)
    for synset in wordnet.synsets(temp):
        for lemma in synset.lemmas():
            sample_words.add(lemma.name())
sample_words = sample_words.union(sample_word)

#print(sample_words)
#print(len(sample_words))

###########################################
#############_Answer Checking_#############
###########################################

for i_cell in range(1, my_row + 1):
    if i_cell<9:
        i_cell = i_cell+1
        cell_obj = my_sheet_obj.cell(row = i_cell, column = 3)

        sheetNumber = cell_obj.value

        third_answer = open( sheetNumber , 'r')
        tex = third_answer.read()

        third_word = keywords(
            tex,
        ).split("\n")

        third_words = set()
        for w in third_word:
            third_words.add(lemmatizer.lemmatize(w))


        #print(third_words)
        #print(len(third_words))

        common_set = third_words.intersection(sample_words)

        ###########################################
        ##############_Grammar Check_##############
        ###########################################

        # Mention the language keyword 
        tool = language_check.LanguageTool('en-US') 
        i = 0
        grammarMarks = 0.4*totalMarks 

        # Path of file which needs to be checked 
        with open(sheetNumber, 'r') as fin: 

            for line in fin: 
                matches = tool.check(line) 
                i = i + len(matches)	 
                pass

        # prints total mistakes which are found from the document 
        #print("No. of mistakes found in document is ", i)

        grammarMarks = grammarMarks - (i*0.25)
        if grammarMarks<0:
            grammarMarks = 0

        ###########################################
        ##################_Result_#################
        ###########################################

        #print(common_set)
        #print(len(common_set))
        #print(len(sample_word))

        contentMarks = 0.6*totalMarks

        contentMatch = (len(common_set)/len(sample_word)) * 50
        contentMatch = int(contentMatch)
        contentMarks = contentMatch * contentMarks / 10

        print("Grammar Marks ", grammarMarks)
        print("Content Marks ", contentMarks)
        obtainedMarks = grammarMarks + contentMarks

        print("You scored ", obtainedMarks, " out of ", totalMarks)

        GM = my_sheet_obj.cell(row = i_cell, column = 4)
        GM.value = grammarMarks

        CM = my_sheet_obj.cell(row = i_cell, column = 5)
        CM.value = contentMarks
        
        c = my_sheet_obj.cell(row = i_cell, column = 6)
        c.value = obtainedMarks

my_wb_obj.save("F:\Projects\Subjective-Q-A-Checker\QA Checker\Result.xlsx")